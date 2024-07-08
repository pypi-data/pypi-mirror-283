from elasticsearch import Elasticsearch
import urllib3, logging
import pandas as pd
import os
from confluent_kafka import Producer
from configparser import ConfigParser
import json
from openpyxl.utils import get_column_letter
import openpyxl
from nltk.translate.bleu_score import sentence_bleu
import numpy as np
import pickle


urllib3.disable_warnings()
logging.captureWarnings(True)


class EsTool:
    def __init__(self, url, id, pwd) -> None:
        self.es = Elasticsearch(
            url,
            http_auth=(id, pwd),
            verify_certs=False,
            timeout=10,
            ssl_show_warn=False,
        )
        self.env = None

    def setup_env(self, env):
        self.env = env

    def make_body(self, **kwargs):
        must_clauses = []
        for key, value in kwargs.items():
            must_clauses.append({"match": {key: value}})
        body = {
            "_source": [],
            "query": {
                "bool": {
                    "must": must_clauses,
                }
            },
            "size": 1000,
            # "sort": [{"UpdateTime": {"order": "desc"}}],
        }
        return body

    def search(self, index_sort, **kwargs):
        if not self.env:
            raise ValueError(
                "환경이 설정되지 않았습니다. setup_env를 먼저 실행해주세요."
            )
        elif self.env == "dev":
            index_name = f"{index_sort}-2024"
        else:
            index_name = f"apply-ai-{self.env}-{index_sort}-2024"
        body = self.make_body(kwargs)
        res = self.es.search(index=index_name, body=body)
        hits = res["hits"]["hits"]
        if hits:
            print("Data Num :", len(hits))
            return hits
        else:
            print("Nodata")


class ExcelTool:
    def __init__(self) -> None:
        self.config = {
            "Comments": {"width": 80},
            "BestSentence1": {"width": 20},
            "BestSentence2": {"width": 20},
            "FeedBack": {"width": 40},
        }

    def setup_config(self, **kwargs):
        self.config.update(kwargs)
        print(f"Config updated to {self.config}")

    def post_excel(self, df: pd.DataFrame, save_path: str):
        writer = pd.ExcelWriter(f"{save_path}.xlsx", engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Sheet1")

        # 엑셀 시트 가져오기
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        # 열 너비 자동 조정 또는 사용자 지정
        for column in df.columns:
            column_width = max(df[column].astype(str).apply(len).max(), len(column))
            if column in self.config:
                if "width" in self.config[column]:
                    column_width = self.config[column]["width"]
            worksheet.column_dimensions[
                get_column_letter(df.columns.get_loc(column) + 1)
            ].width = column_width

        # 첫 번째 행 고정
        worksheet.freeze_panes = "A2"

        # 자동 줄 바꿈 적용
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # 변경 사항 저장하고 파일 닫기
        writer.close()


class FileLoader:
    def __init__(self):
        self.handlers = {
            "txt": self._load_txt,
            "json": self._load_json,
            "csv": self._load_csv,
            "xlsx": self._load_excel,
            "parquet": self._load_parquet,
            "pkl": self._load_pickle,
            "npz": self._load_npz,
            "npy": self._load_npy,
        }

    def load(self, path: str, fn=False):
        try:
            file_name, file_type = os.path.basename(path).split(".")
            file_type = file_type.lower()

            if file_type in self.handlers:
                if not fn:
                    return self.handlers[file_type](path)
                else:
                    return self.handlers[file_type](path), file_name
            else:
                raise ValueError(f"Unsupported file type: {file_type}")

        except Exception as e:
            print(f"An error occurred: {e}")
            return None

    def _load_txt(self, path):
        with open(path, "r", encoding="utf-8") as file:
            return file.read()

    def _load_json(self, path):
        with open(path, "r", encoding="utf-8") as file:
            return json.load(file)

    def _load_csv(self, path):
        return pd.read_csv(path, encoding="utf-8")

    def _load_excel(self, path):
        return pd.read_excel(path)

    def _load_parquet(self, path):
        return pd.read_parquet(path)

    def _load_pickle(self, path):
        with open(path, "rb") as file:
            return pickle.load(file)

    def _load_npz(self, path):
        return np.load(path)

    def _load_npy(self, path):
        return np.load(path)
