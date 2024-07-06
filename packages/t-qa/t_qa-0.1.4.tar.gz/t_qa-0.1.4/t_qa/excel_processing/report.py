"""Module for working with Report."""
import re
import traceback

from openpyxl.workbook import Workbook

from ..columns_blocks import ColumnsBlock, InputsBlock, RecordsBlock, RunDetailsBlock, TestCasesBlock
from ..config import Inputs
from ..excel_processing.google_sheet import GoogleSheet
from ..goole_api.google_drive_service import GoogleDriveService
from ..logger import logger
from ..models import Alignment, Color, RunData, TestCase
from ..status import Status
from ..workitems import VARIABLES


class DumpCell:
    """Dump cell for the Excel file."""

    def __init__(
        self,
        value: str,
        row: int,
        column_number: int,
        color: str = None,
        alignment: str = Alignment.left.value,
    ):
        """Initialize the dump cell."""
        self.row = row
        self.column_number = column_number
        self.value = value
        self.bg_color = color
        self.alignment = alignment


class _Report:
    """QA report."""

    def __init__(self, file_path: str, google_sheet: GoogleSheet = None, google_drive: GoogleDriveService = None):
        """Initialize the Report."""
        # google sheet remote file
        self.google_drive = google_drive
        self.google_sheet = google_sheet
        self.google_sheet_file_id = self.get_file_id()
        # excel local file
        self.local_file_path = file_path
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.workbook.create_sheet("Report")
        self.worksheet = self.workbook["Report"]

    def get_file_id(self) -> str:
        """Get the file ID."""
        folder, file = self.get_folder_name_and_googlesheet_name()
        google_drive_folder = self.google_drive.create_folder_if_not_exists(folder)
        return self.google_drive.create_file_if_not_exists(google_drive_folder, file)["id"]

    def get_folder_name_and_googlesheet_name(self) -> tuple[str, str]:
        """Get the folder name and Google Sheet name."""
        project_admin_code: str = Inputs.ADMIN_CODE
        project_number = re.search(r"([0-9]+)", project_admin_code).group(0)
        folder = project_admin_code.replace(project_number, "").strip().upper()
        file = project_admin_code.strip().upper()
        return folder, file

    def dump(self, run_data: RunData) -> None:
        """Dump the report."""
        headers = self.get_header(run_data=run_data)
        all_cells = self.get_run_data_dump_cells(run_data=run_data, headers=headers)
        try:
            self.google_sheet.check_or_create_headers(file_id=self.google_sheet_file_id)
            self.google_sheet.write_to_google_sheet(all_cells=all_cells, file_id=self.google_sheet_file_id)
        except Exception as e:
            logger.error(f"Error: {e}")
            traceback.print_exc()
            self.write_header_to_excel(headers=headers)
            self.write_cells_to_excel(all_cells=all_cells)

    def write_header_to_excel(self, headers: list[DumpCell]) -> None:
        """Write the header to the Excel file."""
        for cell in headers:
            self.worksheet.cell(row=cell.row, column=cell.column_number, value=cell.value)

    def get_block_header(self, block: ColumnsBlock, block_start_column=0) -> list[DumpCell]:
        """Write the block header."""
        header_block = []
        for index, column_name in enumerate(block.column_names):
            header_block.append(
                DumpCell(
                    value=column_name,
                    row=1,
                    column_number=block_start_column + index,
                    color=block.color,
                    alignment=Alignment.center.value,
                )
            )
        return header_block

    def get_header(self, run_data: RunData) -> list[DumpCell]:
        """Write the header."""
        headers = []
        run_details_start_column = 1
        run_details_block = RunDetailsBlock()
        headers += self.get_block_header(run_details_block, run_details_start_column)

        records_start_column = run_details_start_column + len(run_details_block.column_names)
        records_block = RecordsBlock()
        headers += self.get_block_header(records_block, records_start_column)

        inputs_start_column = records_start_column + len(records_block.column_names)
        inputs_block = InputsBlock()
        headers += self.get_block_header(inputs_block, inputs_start_column)

        test_case_start_column = inputs_start_column + len(inputs_block.column_names)
        test_cases_block = TestCasesBlock(run_data.test_cases)
        headers += self.get_block_header(test_cases_block, test_case_start_column)
        if not self.google_sheet.is_header_exist(self.google_sheet_file_id):
            self.google_sheet.get_header_block(run_details_block, run_details_start_column)
            self.google_sheet.get_header_block(records_block, records_start_column)
            self.google_sheet.get_header_block(inputs_block, inputs_start_column)
            self.google_sheet.get_header_block(test_cases_block, test_case_start_column)
        return headers

    def get_run_data_dump_cells(self, run_data: RunData, headers: list[DumpCell]) -> list[DumpCell]:
        """Dump the run data."""
        row = 2
        all_cells = []
        all_cells += self.get_run_details_cells(run_data, row, headers)
        all_cells += self.get_records_cells(run_data, row, headers)
        all_cells += self.get_inputs_cells(row, headers)
        all_cells += self.get_test_cases_cells(run_data.test_cases, row, headers)
        return all_cells

    def get_run_details_cells(self, run_data: RunData, row: int, headers: list[DumpCell]):
        """Get the run details cells."""
        return [
            DumpCell(
                value=run_data.run_link,
                row=row,
                column_number=self.get_column_number("Run link", headers),
            ),
            DumpCell(
                value=run_data.run_date,
                row=row,
                column_number=self.get_column_number("Date", headers),
            ),
            DumpCell(
                value=run_data.status,
                row=row,
                column_number=self.get_column_number("Status", headers),
            ),
            DumpCell(
                value=run_data.bugs,
                row=row,
                column_number=self.get_column_number("Bugs", headers),
            ),
            DumpCell(
                value=run_data.code_coverage,
                row=row,
                column_number=self.get_column_number("Code Cov", headers),
            ),
            DumpCell(
                value=run_data.empower_env,
                row=row,
                column_number=self.get_column_number("Emp Env", headers),
            ),
            DumpCell(
                value=run_data.duration,
                row=row,
                column_number=self.get_column_number("Duration", headers),
            ),
        ]

    def get_records_cells(self, run_data: RunData, row: int, headers: list[DumpCell]):
        """Get the records cells."""
        return [
            DumpCell(
                value=str(run_data.success_records),
                row=row,
                column_number=self.get_column_number("Success", headers),
            ),
            DumpCell(
                value=str(run_data.total_records),
                row=row,
                column_number=self.get_column_number("Total", headers),
            ),
            DumpCell(
                value=str(run_data.failed_records),
                row=row,
                column_number=self.get_column_number("Failed", headers),
            ),
        ]

    def get_inputs_cells(self, row: int, headers: list[DumpCell]):
        """Get the inputs cells."""
        input_block = InputsBlock().column_names
        cells = []
        for inputs in input_block:
            cells.append(
                DumpCell(
                    value=VARIABLES.get(inputs, "Empty"),
                    row=row,
                    column_number=self.get_column_number(inputs, headers),
                )
            )
        return cells

    def get_test_cases_cells(self, test_cases: list[TestCase], row: int, headers: list[DumpCell]):
        """Get the test cases cells."""
        status_colors = {Status.FAIL.value: Color.red.value, Status.PASS.value: Color.green.value, "": Color.gray.value}
        cells = []
        for test_case in test_cases:
            cells.append(
                DumpCell(
                    value=test_case.status,
                    row=row,
                    column_number=self.get_column_number(test_case.id, headers),
                    color=status_colors[test_case.status],
                    alignment="CENTER",
                )
            )
        return cells

    def get_column_number(self, header_name: str, headers: list[DumpCell]) -> int:
        """Get the column number."""
        for cell in headers:
            if cell.value == header_name:
                return cell.column_number
        else:
            raise ValueError(f"Header {header_name} not found")

    def write_cells_to_excel(self, all_cells: list[DumpCell]) -> None:
        """Write the cells to the Excel file."""
        for cell in all_cells:
            self.worksheet.cell(row=cell.row, column=cell.column_number, value=cell.value)
        self.workbook.save(self.local_file_path)
