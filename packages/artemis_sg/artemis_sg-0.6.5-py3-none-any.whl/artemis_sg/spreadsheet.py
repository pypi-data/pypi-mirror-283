import logging
import math
import os
import re
from copy import copy
from inspect import getsourcefile

from googleapiclient.discovery import build
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from PIL import Image as PIL_Image
from PIL import UnidentifiedImageError

from artemis_sg import app_creds, items, vendor
from artemis_sg.config import CFG

MODULE = os.path.splitext(os.path.basename(__file__))[0]


def get_worksheet(wb_obj, worksheet):
    ws = wb_obj.worksheets[0] if not worksheet else wb_obj[worksheet]
    return ws

def get_sheet_keys(ws):
    for row in ws.values:
        sheet_keys = [x.upper() if isinstance(x, str) else x for x in row]
        break
    return sheet_keys

def shift_col(ws, col_key, target_idx):
    ws.insert_cols(target_idx)
    sheet_keys = get_sheet_keys(ws)
    sheet_key_idx = sheet_keys.index(col_key) + 1  # for openpyxl
    sheet_key_idx_ltr = get_column_letter(sheet_key_idx)
    col_delta = target_idx - sheet_key_idx
    ws.move_range(f"{sheet_key_idx_ltr}1:{sheet_key_idx_ltr}{ws.max_row}",
                  rows=0, cols=col_delta)
    ws.delete_cols(sheet_key_idx)

def copy_cell_style(ws, style_src_cell, target_cell):
    if style_src_cell.has_style:
        ws[target_cell].font = copy(style_src_cell.font)
        ws[target_cell].border = copy(style_src_cell.border)
        ws[target_cell].fill = copy(style_src_cell.fill)
        ws[target_cell].number_format = copy(style_src_cell.number_format)
        ws[target_cell].protection = copy(style_src_cell.protection)
        ws[target_cell].alignment = copy(style_src_cell.alignment)

def freeze_first_row(worksheet):
    logging.info("Freezing first row of worksheet.")
    worksheet.views.sheetView[0].topLeftCell = "A1"
    worksheet.freeze_panes = "A2"

def create_col(ws, col_key, target_idx, style_src_cell=None):
    ws.insert_cols(target_idx)
    col_header = f"{get_column_letter(target_idx)}1"
    ws[col_header] = col_key.title()
    if style_src_cell:
        copy_cell_style(ws, style_src_cell, col_header)

def sequence_worksheet(ws, col_order, isbn_key):
    sheet_keys = get_sheet_keys(ws)
    for i, key_name in enumerate(col_order):
        order_idx = i + 1  # for openpyxl
        if key_name == "ISBN":
            key_name = isbn_key  # noqa: PLW2901
        if key_name in sheet_keys:
            shift_col(ws, key_name, order_idx)
        else:
            create_col(ws, key_name, order_idx)

def size_sheet_cols(ws, isbn_key):
    dim_holder = DimensionHolder(worksheet=ws)
    sheet_keys = get_sheet_keys(ws)
    for i, key_name in enumerate(sheet_keys):
        col_idx = i + 1  # for openpyxl
        col_idx_ltr = get_column_letter(col_idx)
        width = (
            max(len(str(cell.value)) for cell in ws[col_idx_ltr])
            * CFG["asg"]["spreadsheet"]["sheet_image"]["col_buffer"]
        )
        if width > CFG["asg"]["spreadsheet"]["sheet_image"]["max_col_width"]:
            width = CFG["asg"]["spreadsheet"]["sheet_image"]["max_col_width"]
        dim_holder[col_idx_ltr] = ColumnDimension(ws, index=col_idx_ltr, width=width)
        if key_name == isbn_key:
            dim_holder[col_idx_ltr] = ColumnDimension(
                ws,
                index=col_idx_ltr,
                width=math.ceil(
                    CFG["asg"]["spreadsheet"]["sheet_image"]["isbn_col_width"]
                    * CFG["asg"]["spreadsheet"]["sheet_image"]["col_buffer"]
                ),
            )
        if key_name == "IMAGE":
            dim_holder[col_idx_ltr] = ColumnDimension(
                ws,
                index=col_idx_ltr,
                width=CFG["asg"]["spreadsheet"]["sheet_image"]["image_col_width"]
            )

    ws.column_dimensions = dim_holder

def insert_image(image_directory, ws, isbn_cell, image_cell):
    namespace = f"{MODULE}.{insert_image.__name__}"
    image_row_height = CFG["asg"]["spreadsheet"]["sheet_image"]["image_row_height"]
    if isbn_cell.value:
        isbn = isbn_cell.value
        if isinstance(isbn, float):
            isbn = int(isbn)
        elif isinstance(isbn, str):
            m = re.search('="(.*)"', isbn)
            if m:
                isbn = m.group(1)
        try:
            isbn = str(isbn).strip()
        except Exception as e:
            logging.error(f"{namespace}: Err reading isbn '{isbn}', err: '{e}'")
            isbn = ""
    else:
        isbn = ""
    # Set row height
    row_dim = ws.row_dimensions[image_cell.row]
    row_dim.height = image_row_height

    # Insert image into cell
    filename = f"{isbn}.jpg"
    filepath = os.path.join(image_directory, filename)
    logging.debug(f"{namespace}: Attempting to insert '{filepath}'.")
    if os.path.isfile(filepath):
        img = Image(filepath)
        ws.add_image(img, f"{image_cell.column_letter}{image_cell.row}")
        logging.info(f"{namespace}: Inserted '{filepath}'.")


def sheet_image(vendor_code, workbook, worksheet, image_directory, out):
    namespace = f"{MODULE}.{sheet_image.__name__}"

    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")
    vendr = vendor.Vendor(vendor_code)
    vendr.set_vendor_data()

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")

    col_order = CFG["asg"]["spreadsheet"]["sheet_image"]["col_order"]
    sequence_worksheet(ws, col_order, isbn_key)
    size_sheet_cols(ws, isbn_key)

    # Prepare "IMAGE" column
    sk = get_sheet_keys(ws)
    try:
        img_idx = sk.index("IMAGE") + 1
        img_idx_ltr = get_column_letter(img_idx)
    except ValueError as e:
        logging.error(f"{namespace}: Err finding 'IMAGE' column in sheet '{workbook}'.")
        logging.error("Aborting.")
        raise e
    try:
        isbn_idx = sk.index(isbn_key) + 1
        isbn_idx_ltr = get_column_letter(isbn_idx)
    except ValueError as e:
        logging.error(
                f"{namespace}: Err, no '{isbn_key}' column in sheet '{workbook}'.")
        logging.error("Aborting.")
        raise e

    for i in range(1, ws.max_row):
        isbn_cell = ws[f"{isbn_idx_ltr}{i+1}"]
        image_cell = ws[f"{img_idx_ltr}{i+1}"]
        # Format to center content
        image_cell.alignment = Alignment(horizontal="center")
        insert_image(image_directory, ws, isbn_cell, image_cell)

    freeze_first_row(ws)

    # Save workbook
    wb.save(out)

def validate_isbn(isbn):
    namespace = f"{MODULE}.{validate_isbn.__name__}"
    valid_isbn = ""
    if isinstance(isbn, str):
        m = re.search('="(.*)"', isbn)
        if m:
            isbn = m.group(1)
    try:
        valid_isbn = str(int(isbn)).strip()
    except Exception as e:
        logging.error(f"{namespace}: Err reading isbn '{isbn}', err: '{e}'")
        valid_isbn = ""
    return valid_isbn


def validate_qty(qty):
    namespace = f"{MODULE}.{validate_qty.__name__}"
    try:
        valid_qty = str(int(qty)).strip()
    except Exception as e:
        logging.error(f"{namespace}: Err reading Order qty '{qty}', err: '{e}'")
        valid_qty = None
    return valid_qty


def get_order_items(vendor_code, workbook, worksheet):
    namespace = f"{MODULE}.{get_order_items.__name__}"

    order_items = []
    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")
    vendr = vendor.Vendor(vendor_code)
    vendr.set_vendor_data()

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")

    # Find Isbn and Order column letters
    row01 = ws[1]
    for cell in row01:
        if cell.value == isbn_key:
            isbn_column_letter = cell.column_letter
        if cell.value == "Order":
            order_column_letter = cell.column_letter

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter == isbn_column_letter:
                isbn_cell = cell
            if cell.column_letter == order_column_letter:
                order_cell = cell
        # Validate ISBN
        isbn = validate_isbn(isbn_cell.value)
        if not isbn:
            continue
        # Validate Order Qty
        qty = validate_qty(order_cell.value)
        if not qty:
            continue
        order_items.append((isbn, qty))

    return order_items


def mkthumbs(image_directory):
    namespace = f"{MODULE}.{mkthumbs.__name__}"

    thumb_width = CFG["asg"]["spreadsheet"]["mkthumbs"]["width"]
    thumb_height = CFG["asg"]["spreadsheet"]["mkthumbs"]["height"]

    here = os.path.dirname(getsourcefile(lambda: 0))
    data = os.path.abspath(os.path.join(here, "data"))
    logo = os.path.join(data, "artemis_logo.png")
    logging.debug(f"{namespace}: Found image for thumbnail background at '{logo}'")
    sub_dir = "thumbnails"
    back = PIL_Image.open(logo)
    thumb_dir = os.path.join(image_directory, sub_dir)
    logging.debug(f"{namespace}: Defining thumbnail directory as '{thumb_dir}'")
    if not os.path.isdir(thumb_dir):
        logging.debug(f"{namespace}: Creating directory '{thumb_dir}'")
        os.mkdir(thumb_dir)
        if os.path.isdir(thumb_dir):
            logging.info(f"{namespace}: Successfully created directory '{thumb_dir}'")
        else:
            logging.error(
                f"{namespace}: Failed to create directory '{thumb_dir}'. Aborting."
            )
            raise Exception
    files = os.listdir(image_directory)
    for f in files:
        # Valid files are JPG or PNG that are not supplemental images.
        image = re.match(r"^.+\.(?:jpg|png)$", f)
        if not image:
            continue
        # Supplemental images have a "-[0-9]+" suffix before the file type.
        # AND a file without that suffix exists int he image_directory.
        suffix = re.match(r"(^.+)-[0-9]+(\.(?:jpg|png))$", f)
        if suffix:
            primary = suffix.group(1) + suffix.group(2)
            primary_path = os.path.join(image_directory, primary)
            if os.path.isfile(primary_path):
                continue
        thumb_file = os.path.join(thumb_dir, f)
        # don't remake thumbnails
        if os.path.isfile(thumb_file):
            continue
        bk = back.copy()
        try:
            file_path = os.path.join(image_directory, f)
            fg = PIL_Image.open(file_path)
        except UnidentifiedImageError:
            logging.error(f"{namespace}: Err reading '{f}', deleting '{file_path}'")
            os.remove(file_path)
            continue
        fg.thumbnail((thumb_width, thumb_height))
        size = (int((bk.size[0] - fg.size[0]) / 2), int((bk.size[1] - fg.size[1]) / 2))
        bk.paste(fg, size)
        logging.debug(f"{namespace}: Attempting to save thumbnail '{thumb_file}'")
        bkn = bk.convert("RGB")
        bkn.save(thumb_file)
        logging.info(f"{namespace}: Successfully created thumbnail '{thumb_file}'")


def get_sheet_data(workbook, worksheet=None):
    namespace = f"{MODULE}.{get_sheet_data.__name__}"
    #########################################################################
    # Try to open sheet_id as an Excel file
    sheet_data = []
    try:
        wb = load_workbook(workbook)
        ws = get_worksheet(wb, worksheet)
        for row in ws.values:
            sheet_data.append(row)
    except (FileNotFoundError, InvalidFileException):
        #########################################################################
        # Google specific stuff
        # authenticate to google sheets
        logging.info(f"{namespace}: Authenticating to google api.")
        creds = app_creds.app_creds()
        sheets_api = build("sheets", "v4", credentials=creds)
        # get sheet data
        if not worksheet:
            sheets = (
                sheets_api.spreadsheets()
                .get(spreadsheetId=workbook)
                .execute()
                .get("sheets", "")
            )
            ws = sheets.pop(0).get("properties", {}).get("title")
        else:
            ws = worksheet
        sheet_data = (
            sheets_api.spreadsheets()
            .values()
            .get(range=ws, spreadsheetId=workbook)
            .execute()
            .get("values")
        )
        #########################################################################
    return sheet_data


def sheet_waves(vendor_code, workbook, worksheet, out, scraped_items_db):
    namespace = f"{MODULE}.{sheet_waves.__name__}"

    addl_data_columns = [
            "Description",
            "Dimension",
            ]
    addl_image_columns = [
            "ImageURL0",
            "ImageURL1",
            "ImageURL2",
            "ImageURL3",
            "ImageURL4",
            "ImageURL5",
            "ImageURL6",
            ]
    addl_columns = addl_data_columns + addl_image_columns
    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")
    vendr = vendor.Vendor(vendor_code)
    vendr.set_vendor_data()

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    sheet_data = get_sheet_data(workbook, worksheet)

    sheet_keys = [x for x in sheet_data.pop(0) if x]  # filter out None
    items_obj = items.Items(sheet_keys, sheet_data, vendr.isbn_key)
    items_obj.load_scraped_data(scraped_items_db)

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")

    # Append columns
    col_insert_idx = ws.max_column + 1
    ws.insert_cols(col_insert_idx, len(addl_columns))
    i = 1
    for col in addl_columns:
        col_idx = col_insert_idx + i
        ws.cell(row=1, column=col_idx, value=col)
        i = i + 1

    # Find ISBN column
    row01 = ws[1]
    isbn_idx = None
    for cell in row01:
        if isinstance(cell.value, str) and cell.value.upper() == isbn_key.upper():
            isbn_idx = cell.column - 1
            break
    if isbn_idx is None:
        logging.error(f"{namespace}: Err no isbn column in spreadsheet")
        raise Exception

    # Insert data in cells
    for row in ws.iter_rows(min_row=2):
        # get isbn cell
        isbn = str(row[isbn_idx].value)
        # find items_obj matching isbn
        item = items_obj.find_item(isbn)
        if item:
            idx = col_insert_idx
            for key in addl_data_columns:
                if key.upper() in item.data:
                    row[idx].value = item.data[key.upper()]
                idx = idx + 1
            for img_url in item.image_urls[:7]:
                row[idx].value = img_url
                idx = idx + 1

    freeze_first_row(ws)

    # Save workbook
    wb.save(out)
