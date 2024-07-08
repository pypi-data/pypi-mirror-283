import os

import matplotlib.pyplot as plt
import scipy.stats as stats
import seaborn as sns
import starbars
from openpyxl import load_workbook

__all__ = [
    "value_graph_lipid_class",
    "value_graph_region",
    "values_graph_lipid",
    "zscore_graph_lipid_class",
    "zscore_graph_region",
    "zscore_graph_lipid",
]


def save_sheet(comment, sheet_name, output_path):

    wb = load_workbook(output_path + "/output/Output file.xlsx")
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]
        ws.append(comment)
        wb.save(output_path + "/output/Output file.xlsx")
    else:
        ws = wb[sheet_name]
        ws.append(comment)
        wb.save(output_path + "/output/Output file.xlsx")


def get_test(shapiro, levene, control_values, experimental_values):
    if shapiro < 0.05 and levene < 0.05:
        stat, pvalue = stats.ttest_ind(control_values, experimental_values)
        test = "T-Test"
    elif shapiro < 0.05 and levene > 0.05:
        stat, pvalue = stats.ttest_ind(control_values, experimental_values, equal_var=False)
        test = "Welch T-Test"
    elif shapiro > 0.05 and levene > 0.05:
        stat, pvalue = stats.mannwhitneyu(control_values, experimental_values)
        test = "Mann Whitney"
    else:
        pvalue = 0
        test = "no test"

    return pvalue, test


# Graphs by Z scores
def zscore_graph_lipid(df_final, control_name, experimental_name, output_path, palette, show=True):

    if not os.path.exists(output_path + "/output/zscore_graphs/lipid"):
        os.makedirs(output_path + "/output/zscore_graphs/lipid")
    order = [control_name, experimental_name]

    for (region, lipid), data in df_final.groupby(["Regions", "Lipids"]):
        shapiro = data.iloc[0]["Shapiro Normality"]
        levene = data.iloc[0]["Levene Equality"]
        control_group = data[data["Genotype"] == control_name]
        experimental_group = data[data["Genotype"] != control_name]
        control_values = control_group["Z Scores"]
        experimental_values = experimental_group["Z Scores"]

        print(f"Creating graph for {lipid} in {region}")

        sns.boxplot(x="Genotype", y="Z Scores", data=data, order=order, hue="Genotype", palette=palette)
        sns.stripplot(x="Genotype", y="Z Scores", data=data, order=order, color="k", size=4)

        pvalue, test = get_test(shapiro, levene, control_values, experimental_values)
        pairs = [(control_name, experimental_name, pvalue)]
        plt.xlabel("Genotype")
        plt.ylabel("Z Scores")
        plt.title(f"Z Scores Distribution for {lipid} in {region}: {control_name} vs {experimental_name}")
        starbars.draw_annotation(pairs)
        plt.savefig(output_path + "/output/graphs/" + f"Z Scores for {lipid} in {region}.png", dpi=1200)
        if show:
            plt.show()
        plt.close()

        comment = [f"For z scores of {lipid} in {region}, {test} was performed. P-value is {pvalue}."]
        save_sheet(comment, "Comments", output_path)


def zscore_graph_region(df_final, control_name, experimental_name, output_path, palette, show=True):

    if not os.path.exists(output_path + "/output/zscore_graphs/region"):
        os.makedirs(output_path + "/output/zscore_graphs/region")

    for region, data in df_final.groupby("Regions"):
        print(f"Creating graph for: {region}")

        sns.boxplot(x="Lipids", y="Z Scores", hue="Genotype", data=data, palette=palette)
        sns.stripplot(x="Lipids", y="Z Scores", data=data, hue="Genotype", dodge=True, color="k", size=4)

        plt.xlabel("Lipids")
        plt.ylabel("Z Scores")
        plt.title(f"Z Scores Distribution in {region}: {control_name} vs {experimental_name}")
        plt.xticks(rotation=90)

        plt.savefig(output_path + f"/output/zscore_graphs/region/Z-Scores Distribution {region}.png", dpi=1200)
        if show:
            plt.show()
        plt.close()


def zscore_graph_lipid_class(df_final, control_name, experimental_name, output_path, palette, show=True):

    if not os.path.exists(output_path + "/output/zscore_graphs/lipid_class"):
        os.makedirs(output_path + "/output/zscore_graphs/lipid_class")

    for region, data in df_final.groupby("Regions"):
        print(f"Creating lipid classes' graph for {region}")

        sns.boxplot(
            x="Lipid Class",
            y="Z Scores",
            hue="Genotype",
            data=data,
            order=data["Lipid Class"].unique(),
            palette=palette,
        )
        sns.stripplot(x="Lipid Class", y="Z Scores", data=data, dodge=True, color="k", size=4)

        plt.xlabel("Lipid Class")
        plt.ylabel("Z Score")
        plt.title(f"Z Scores Distribution in {region}: {control_name} vs {experimental_name}")
        plt.xticks(rotation=90)

        plt.savefig(output_path + f"/output/zscore_graphs/lipid_class/Z-Scores Distribution {region}.png", dpi=1200)
        if show:
            plt.show()
        plt.close()


# Graphs by values
def values_graph_lipid(df_final, control_name, experimental_name, output_path, palette, show=True):

    if not os.path.exists(output_path + "/output/value_graphs/lipid"):
        os.makedirs(output_path + "/output/value_graphs/lipid")
    order = [control_name, experimental_name]

    for (region, lipid), data in df_final.groupby(["Regions", "Lipids"]):
        shapiro = data.iloc[0]["Shapiro Normality"]
        levene = data.iloc[0]["Levene Equality"]
        control_group = data[data["Genotype"] == control_name]
        experimental_group = data[data["Genotype"] != control_name]
        control_values = control_group["Normalized Values"]
        experimental_values = experimental_group["Normalized Values"]

        print(f"Creating graph for {lipid} in {region}")

        sns.boxplot(x="Genotype", y="Normalized Values", data=data, order=order, hue="Genotype", palette=palette)
        sns.stripplot(x="Genotype", y="Normalized Values", data=data, order=order, color="k", size=4)

        pvalue, test = get_test(shapiro, levene, control_values, experimental_values)
        pairs = [(control_name, experimental_name, pvalue)]
        plt.xlabel("Genotype")
        plt.ylabel("Normalized Values")
        plt.title(f"Normalized Values Distribution for {lipid} in {region}: {control_name} vs {experimental_name}")
        starbars.draw_annotation(pairs)
        plt.savefig(output_path + f"/output/value_graphs/lipid/Normalized Values for {lipid} in {region}.png", dpi=1200)
        if show:
            plt.show()
        plt.close()

        comment = [f"For values of {lipid} in {region}, {test} was performed. P-value is {pvalue}."]
        save_sheet(comment, "Comments", output_path)


def value_graph_region(df_final, control_name, experimental_name, output_path, palette, show=True):

    if not os.path.exists(output_path + "/output/value_graphs/region"):
        os.makedirs(output_path + "/output/value_graphs/region")

    for region, data in df_final.groupby("Regions"):

        print(f"Creating graph for {region}")

        plt.figure(figsize=(16, 13))

        sns.boxplot(x="Lipids", y="Normalized Values", hue="Genotype", data=data, palette=palette)
        sns.stripplot(x="Lipids", y="Normalized Values", data=data, hue="Genotype", dodge=True, color="k", size=4)

        plt.xlabel("Lipids")
        plt.ylabel("Normalized Values")
        plt.title(f"Normalized Values Distribution in {region}: {control_name} vs {experimental_name}")
        plt.xticks(rotation=90)

        plt.savefig(output_path + f"/output/value_graphs/region/Normalized Values Distribution {region}.png", dpi=1200)
        if show:
            plt.show()
        plt.close()


def value_graph_lipid_class(df_final, control_name, experimental_name, output_path, palette, show=True):

    if not os.path.exists(output_path + "/output/value_graphs/lipid_class"):
        os.makedirs(output_path + "/output/value_graphs/lipid_class")

    for region, data in df_final.groupby("Regions"):
        print(f"Creating lipid classes' graph for {region}")

        sns.boxplot(
            x="Lipid Class",
            y="Normalized Values",
            hue="Genotype",
            data=data,
            order=data["Lipid Class"].unique(),
            palette=palette,
        )
        sns.stripplot(x="Lipid Class", y="Normalized Values", data=data, dodge=True, color="k", size=4)

        plt.xlabel("Lipid Class")
        plt.ylabel("Normalized Values")
        plt.title(f"Normalized Values Distribution in {region}: {control_name} vs {experimental_name}")
        plt.xticks(rotation=90)

        plt.savefig(
            output_path + f"/output/value_graphs/lipid_class/Normalized Values Distribution {region}.png", dpi=1200
        )
        if show:
            plt.show()
        plt.close()
