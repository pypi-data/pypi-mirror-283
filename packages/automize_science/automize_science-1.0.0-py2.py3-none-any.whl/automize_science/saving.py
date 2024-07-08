import itertools

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def color_df(df_final, sheet, output_path):
    wb = load_workbook(output_path + "/output/Output file.xlsx")
    ws = wb[sheet]

    # Generate unique region names
    unique_regions = df_final["Regions"].unique()

    # Generate a list of colors (repeat colors if there are more regions than colors)
    colors = ["EA9EBF", "D0A0EC", "B4DFFF", "FFEEBC", "DCEDC1"]
    color_cycle = itertools.cycle(colors)

    # Create a dictionary to map regions to colors
    region_fill_colors = {
        region: PatternFill(start_color=color, end_color=color, fill_type="solid")
        for region, color in zip(unique_regions, color_cycle)
    }

    # Iterate through rows and apply fill colors based on 'Regions' column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        region_value = row[0].value
        region_fill = region_fill_colors.get(region_value, None)

        if region_fill:
            for cell in row:
                cell.fill = region_fill

    wb.save(output_path + "/output/Output file.xlsx")


def save_values(df_final, output_path):
    df_save = df_final.pivot_table(
        index=["Regions", "Mouse ID", "Genotype"],
        columns=["Lipids", "Lipid Class"],
        values=["Values", "Normalized Values"],
        aggfunc="first",
    )
    df_save.reset_index(inplace=True)

    # Save the eliminated lipids and the normalized data with the Z Scores
    with pd.ExcelWriter(output_path + "/output/Output file.xlsx", engine="openpyxl", mode="a") as writer:
        df_save.to_excel(writer, sheet_name="Data for Correlations")

    color_df(df_final, sheet="Data for Correlations", output_path=output_path)


def save_zscores(df_final, output_path):
    df_save = df_final.pivot_table(
        index=["Regions", "Mouse ID", "Genotype"],
        columns=["Lipids", "Lipid Class"],
        values=["Average Z Scores", "Z Scores"],
    )
    df_save.reset_index(inplace=True)

    # Save the eliminated lipids and the normalized data with the Z Scores
    with pd.ExcelWriter(output_path + "/output/Output file.xlsx", engine="openpyxl", mode="a") as writer:
        df_save.to_excel(writer, sheet_name="Z Scores")

    color_df(df_final, sheet="Z Scores", output_path=output_path)

    print("Saving to output file")
